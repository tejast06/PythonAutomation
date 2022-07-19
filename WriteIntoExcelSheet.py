import openpyxl

#same data in rows and column
# file="C:/Users/Hp/Desktop/pythontable.xlsx"
# workbook=openpyxl.load_workbook(file)
# sheet=workbook["Sheet2"]
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="welcome"  # for writing the value in cell
#
# workbook.save(file)

file="C:/Users/Hp/Desktop/pythontable.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet2"]


sheet.cell(1,1).value="tejas"  # for writing the value in cell
sheet.cell(1,2).value="toley"
workbook.save(file)

